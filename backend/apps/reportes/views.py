from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from django.db.models.functions import TruncMonth
from django.db.models import Count, F
from django.http import HttpResponse
from django.utils import timezone
from apps.eventos.models import Evento, CategoriaEvento
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def eventos_por_mes(request):
    """
    Retorna la cantidad de eventos agrupados por mes.
    Query params opcionales: inicio (YYYY-MM-DD), fin (YYYY-MM-DD)
    """

    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden acceder a los reportes.")
    
    qs = Evento.objects.all()
    
    # Filtros opcionales de fecha
    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    
    if inicio:
        try:
            qs = qs.filter(fecha_inicio__date__gte=inicio)
        except ValueError:
            return Response({'error': 'Formato de fecha inicio inválido'}, status=400)
    
    if fin:
        try:
            qs = qs.filter(fecha_inicio__date__lte=fin)
        except ValueError:
            return Response({'error': 'Formato de fecha fin inválido'}, status=400)

    # Agrupar por mes y contar
    datos = (
        qs.annotate(mes=TruncMonth('fecha_inicio'))
          .values('mes')
          .annotate(total=Count('id'))
          .order_by('mes')
    )
    
    out = [
        {
            'mes': d['mes'].isoformat() if d['mes'] else None, 
            'total': d['total']
        } 
        for d in datos
    ]
    
    return Response(out)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def eventos_por_usuario(request):
    """
    Retorna la cantidad de eventos creados por cada organizador.
    """

    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden acceder a los reportes.")
    
    qs = (
        Evento.objects
        .values('organizador__id', 'organizador__username')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    out = [
        {
            'usuario_id': r['organizador__id'], 
            'username': r['organizador__username'] or 'Usuario eliminado', 
            'total': r['total']
        } 
        for r in qs
    ]
    
    return Response(out)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def eventos_por_categoria(request):
    """
    Retorna la cantidad de eventos por categoría.
    """

    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden acceder a los reportes.")
    
    qs = (
        Evento.objects
        .values('categoria__id', 'categoria__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    out = [
        {
            'categoria_id': r['categoria__id'], 
            'nombre': r['categoria__nombre'] or 'Sin categoría', 
            'total': r['total']
        } 
        for r in qs
    ]
    
    return Response(out)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def eventos_por_lugar(request):
    """
    Retorna la cantidad de eventos por ubicación.
    """
    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden acceder a los reportes.")
    
    qs = (
        Evento.objects
        .values('ubicacion')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    
    out = [
        {
            'ubicacion': r['ubicacion'], 
            'total': r['total']
        } 
        for r in qs
    ]
    
    return Response(out)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def eventos_por_estado(request):
    """
    Retorna la cantidad de eventos según su estado:
    - futuros: aún no han comenzado
    - en_curso: están ocurriendo ahora
    - finalizados: ya terminaron
    - llenos: han alcanzado el aforo máximo
    """
    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden acceder a los reportes.")
    
    ahora = timezone.now()
    
    futuros = Evento.objects.filter(fecha_inicio__gt=ahora).count()
    en_curso = Evento.objects.filter(
        fecha_inicio__lte=ahora, 
        fecha_fin__gte=ahora
    ).count()
    finalizados = Evento.objects.filter(fecha_fin__lt=ahora).count()
    
    # Eventos llenos (inscripciones >= aforo)
    llenos = (
        Evento.objects
        .annotate(insc=Count('inscripciones'))
        .filter(insc__gte=F('aforo'))
        .count()
    )
    
    return Response({
        'futuros': futuros,
        'en_curso': en_curso,
        'finalizados': finalizados,
        'llenos': llenos
    })

def generar_csv_response(filename, headers, rows):
    """
    Genera una respuesta HTTP con un archivo CSV.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    
    for row in rows:
        writer.writerow([row.get(h, '') for h in headers])
    
    response = HttpResponse(
        buffer.getvalue(), 
        content_type='text/csv; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_csv(request):
    """
    Exporta datos a CSV según el tipo especificado.
    Query params: tipo (mes, usuarios, categorias, lugares, global)
    """

    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden exportar reportes.")
    
    tipo = request.GET.get('tipo', 'global')
    
    if tipo == 'mes':
        datos = list(
            Evento.objects
            .annotate(mes=TruncMonth('fecha_inicio'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        rows = [
            {
                'mes': d['mes'].strftime('%Y-%m') if d['mes'] else '', 
                'total': d['total']
            } 
            for d in datos
        ]
        return generar_csv_response(
            'eventos_por_mes.csv', 
            ['mes', 'total'], 
            rows
        )
    
    elif tipo == 'usuarios':
        qs = (
            Evento.objects
            .values('organizador__username')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        rows = [
            {
                'username': r['organizador__username'] or 'Usuario eliminado', 
                'total': r['total']
            } 
            for r in qs
        ]
        return generar_csv_response(
            'eventos_por_usuario.csv', 
            ['username', 'total'], 
            rows
        )
    
    elif tipo == 'categorias':
        qs = (
            Evento.objects
            .values('categoria__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        rows = [
            {
                'categoria': r['categoria__nombre'] or 'Sin categoría', 
                'total': r['total']
            } 
            for r in qs
        ]
        return generar_csv_response(
            'eventos_por_categoria.csv', 
            ['categoria', 'total'], 
            rows
        )
    
    elif tipo == 'lugares':
        qs = (
            Evento.objects
            .values('ubicacion')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        rows = [
            {
                'ubicacion': r['ubicacion'], 
                'total': r['total']
            } 
            for r in qs
        ]
        return generar_csv_response(
            'eventos_por_lugar.csv', 
            ['ubicacion', 'total'], 
            rows
        )
    
    else:  # global
        qs = Evento.objects.all().values(
            'id', 'titulo', 'fecha_inicio', 'fecha_fin', 
            'ubicacion', 'aforo', 'organizador__username', 
            'categoria__nombre'
        )
        rows = [
            {
                'id': r['id'],
                'titulo': r['titulo'],
                'organizador': r['organizador__username'],
                'categoria': r['categoria__nombre'] or 'Sin categoría',
                'fecha_inicio': r['fecha_inicio'].strftime('%Y-%m-%d %H:%M'),
                'fecha_fin': r['fecha_fin'].strftime('%Y-%m-%d %H:%M'),
                'ubicacion': r['ubicacion'],
                'aforo': r['aforo']
            } 
            for r in qs
        ]
        return generar_csv_response(
            'eventos_global.csv', 
            ['id', 'titulo', 'organizador', 'categoria', 
             'fecha_inicio', 'fecha_fin', 'ubicacion', 'aforo'], 
            rows
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_xlsx(request):
    """
    Exporta datos a Excel (XLSX) según el tipo especificado.
    Query params: tipo (mes, usuarios, categorias, lugares, global)
    """
    # Verificar que el usuario sea staff
    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden exportar reportes.")
    
    tipo = request.GET.get('tipo', 'global')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if tipo == 'mes':
        datos = list(
            Evento.objects
            .annotate(mes=TruncMonth('fecha_inicio'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        
        ws.append(['Mes', 'Total de Eventos'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for d in datos:
            ws.append([
                d['mes'].strftime('%Y-%m') if d['mes'] else '', 
                d['total']
            ])
    
    elif tipo == 'usuarios':
        qs = (
            Evento.objects
            .values('organizador__username')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        ws.append(['Usuario', 'Total de Eventos'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for r in qs:
            ws.append([
                r['organizador__username'] or 'Usuario eliminado', 
                r['total']
            ])
    
    elif tipo == 'categorias':
        qs = (
            Evento.objects
            .values('categoria__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        ws.append(['Categoría', 'Total de Eventos'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for r in qs:
            ws.append([
                r['categoria__nombre'] or 'Sin categoría', 
                r['total']
            ])
    
    elif tipo == 'lugares':
        qs = (
            Evento.objects
            .values('ubicacion')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        ws.append(['Ubicación', 'Total de Eventos'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for r in qs:
            ws.append([r['ubicacion'], r['total']])
    
    else:  # global
        qs = Evento.objects.all().values(
            'id', 'titulo', 'organizador__username', 'categoria__nombre',
            'fecha_inicio', 'fecha_fin', 'ubicacion', 'aforo'
        )
        
        ws.append([
            'ID', 'Título', 'Organizador', 'Categoría', 
            'Fecha Inicio', 'Fecha Fin', 'Ubicación', 'Aforo'
        ])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for r in qs:
            ws.append([
                r['id'],
                r['titulo'],
                r['organizador__username'] or 'Usuario eliminado',
                r['categoria__nombre'] or 'Sin categoría',
                r['fecha_inicio'].strftime('%Y-%m-%d %H:%M'),
                r['fecha_fin'].strftime('%Y-%m-%d %H:%M'),
                r['ubicacion'],
                r['aforo']
            ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    response = HttpResponse(
        stream.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}.xlsx"'
    
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_pdf(request):
    """
    Exporta datos a PDF según el tipo especificado.
    Query params: tipo (mes, usuarios, categorias, lugares, global)
    """
    # Verificar que el usuario sea staff
    if not request.user.is_staff:
        raise PermissionDenied("Solo administradores pueden exportar reportes.")
    
    tipo = request.GET.get('tipo', 'global')
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a56db'),
        spaceAfter=30,
        alignment=1  # centrado
    )
    
    # Selección de datos según tipo
    if tipo == 'mes':
        datos = list(
            Evento.objects
            .annotate(mes=TruncMonth('fecha_inicio'))
            .values('mes')
            .annotate(total=Count('id'))
            .order_by('mes')
        )
        
        title = Paragraph("Reporte: Eventos por Mes", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        table_data = [['Mes', 'Total de Eventos']]
        for d in datos:
            table_data.append([
                d['mes'].strftime('%Y-%m') if d['mes'] else '',
                str(d['total'])
            ])
    
    elif tipo == 'usuarios':
        qs = list(
            Evento.objects
            .values('organizador__username')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        title = Paragraph("Reporte: Eventos por Usuario", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        table_data = [['Usuario', 'Total de Eventos']]
        for r in qs:
            table_data.append([
                r['organizador__username'] or 'Usuario eliminado',
                str(r['total'])
            ])
    
    elif tipo == 'categorias':
        qs = list(
            Evento.objects
            .values('categoria__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        title = Paragraph("Reporte: Eventos por Categoría", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        table_data = [['Categoría', 'Total de Eventos']]
        for r in qs:
            table_data.append([
                r['categoria__nombre'] or 'Sin categoría',
                str(r['total'])
            ])
    
    elif tipo == 'lugares':
        qs = list(
            Evento.objects
            .values('ubicacion')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        
        title = Paragraph("Reporte: Eventos por Ubicación", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        table_data = [['Ubicación', 'Total de Eventos']]
        for r in qs:
            table_data.append([r['ubicacion'], str(r['total'])])
    
    else:  # global
        qs = list(
            Evento.objects.all().values(
                'id', 'titulo', 'organizador__username', 'categoria__nombre',
                'fecha_inicio', 'fecha_fin', 'ubicacion', 'aforo'
            )
        )
        
        title = Paragraph("Reporte: Listado General de Eventos", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        table_data = [['ID', 'Título', 'Organizador', 'Categoría', 'Inicio', 'Fin', 'Ubicación', 'Aforo']]
        for r in qs:
            table_data.append([
                str(r['id']),
                r['titulo'][:30] + '...' if len(r['titulo']) > 30 else r['titulo'],
                (r['organizador__username'] or 'N/A')[:15],
                (r['categoria__nombre'] or 'N/A')[:15],
                r['fecha_inicio'].strftime('%d/%m/%Y'),
                r['fecha_fin'].strftime('%d/%m/%Y'),
                r['ubicacion'][:20] + '...' if len(r['ubicacion']) > 20 else r['ubicacion'],
                str(r['aforo'])
            ])
    
    # Crear tabla
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(table)
    
    # Pie de página con fecha
    elements.append(Spacer(1, 30))
    footer_text = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}.pdf"'
    
    return response